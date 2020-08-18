# import the package
import openpyxl

# load the sheets with path to document
book = openpyxl.load_workbook("FinancialSample.xlsx")
# get active sheet
sheets = book.active

# get desire cell
cell = sheets.cell(row=1, column=2)
# extract the value
print(cell.value)
sheets.cell(row=2, column=2).value = "Bulgaria"
bg = sheets.cell(row=2, column=2)
print(bg.value)

print(sheets.max_row)
print(sheets.max_column)
# other way to print a cell
print(sheets["B6"].value)
dick = {}
for i in range(1, 8):
    # if sheets.cell(row=i, column=2).value == "Germany":
    #    continue
    for j in range(2, 5):
        dick[sheets.cell(row=1, column=j).value] = sheets.cell(row=2, column=j).value
        print(sheets.cell(row=i, column=j).value, end=" ")
    print()
print(dick)

